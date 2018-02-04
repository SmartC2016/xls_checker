#!/Library/Frameworks/Python.framework/Versions/3.6/bin/python3
# -*- coding: utf8 -*-


"""
Dieses Programm durchsucht mehrere in einer Datei angegebenen Dateien nach definierten Fehlern
Die Dateien können als xlsb angegeben werden und werden dann für die Untersuchung temporär 
in xlsx umgewandelt. Die Fehler werden sowohl als Übersicht angegeben, als auch im Detail in
einer Excel-Datei gespeichert.

Am 27.01.18 als GUI neu aufgesetzt.
"""

__version__ = "0.99b - 04.02.2018"
__author__ = "Christian Hetmann"

# todo xlsb-Datei mit Excel einlesen, in xlsx konvertieren und dann speichern

import tkinter as tk
from tkinter import ttk
from tkinter import scrolledtext
from tkinter.filedialog import askopenfilename
from tkinter import messagebox
import os
import sys
import time
import locale

LARGE_FONT = ("Verdana", 12)
# Breite der Buttons
BWIDTH = 21

class Klaus_App(tk.Frame):
    def __init__(self, parent, *args, **kwargs):
        tk.Frame.__init__(self, parent, *args, **kwargs)
        self.fenster = parent
        self.fenster.title("Klaus' Excel-Checker")

        # Bildschirmbreite ermitteln und Fensterposition bestimmen
        x_pos = int((self.fenster.winfo_screenwidth() - self.fenster.winfo_reqwidth()) / 3)
        y_pos = int((self.fenster.winfo_screenheight() - self.fenster.winfo_reqheight()) / 3)
        self.fenster.geometry(f'{x_pos}x{2*y_pos}+{x_pos}+{y_pos}')

        # Überschreibt die Betriebssystem Funktion, wenn auf das 'x' zum schliessen gedrückt wird.
        # So kann noch was anderes vorher geprüft / gemacht werden, bevor das Programm endet.
        self.fenster.protocol('WM_DELETE_WINDOW', self.click_beenden)

        # Allgemeine Variablen definieren
        self.dateiliste = []  # Hier werden später alle Datei-Namen gespeichert
        self.fehler = False
        self.gespeichert = True
        self.Liste_Fehler = []

        # Tab-Control erstellen
        self.tab_control = ttk.Notebook(self.fenster)
        self.tab_control.grid(row=0, column=0, columnspan=50, rowspan=50, sticky='NESW')

        # Tabs erstellen
        self.tab1 = ttk.Frame(self.tab_control)
        self.tab2 = ttk.Frame(self.tab_control)
        self.tab3 = ttk.Frame(self.tab_control)
        self.tab4 = ttk.Frame(self.tab_control)

        # Beim "Grid" Window Manager sollte man einstellen, wieviel Spalten und Zeilen das Fenster hat
        tabs = [self.tab1, self.tab2, self.tab3, self.tab4]
        zeile = 0
        while zeile < 30:
            for t in tabs:
                t.rowconfigure(zeile, weight=1)
                t.columnconfigure(zeile, weight=1)
            zeile += 1

        # Tabs benennen
        self.tab_control.add(self.tab1, text='Status')
        self.tab_control.add(self.tab2, text='Zusammenfassung')
        self.tab_control.add(self.tab3, text='Details')
        self.tab_control.add(self.tab4, text='Dateiliste')

        # Tab 1 - Status
        # todo Buttons erstellen
        self.label1 = ttk.Label(self.tab1, text='STATUS')
        self.label1.grid(row=0, column=0, sticky='NSEW')

        self.tb1 = scrolledtext.ScrolledText(self.tab1, wrap=tk.WORD, height=25)
        self.tb1.grid(row=1, column=0, columnspan=39, sticky='NSEW')

        self.btn_excel = ttk.Button(self.tab1, text='Excel-Check starten',
                                    default='active', command=self.excel_focus, width=BWIDTH)
        self.btn_excel.grid(row=29, column=0, padx=2, pady=2)

        self.btn_dl_lesen = ttk.Button(self.tab1, text='Dateiliste neu lesen',
                                       command=self.existiert_dateiliste, width=BWIDTH)
        self.btn_dl_lesen.grid(row=29, column=15, padx=2, pady=2)

        self.btn_quit_t1 = ttk.Button(self.tab1, text='Beenden', command=self.click_beenden, width=BWIDTH)
        self.btn_quit_t1.grid(row=29, column=30, padx=2, pady=2)

        # Tab 2 - Zusammenfassung
        # todo checken ob files xlsb oder xlsx sind
        self.label2 = ttk.Label(self.tab2, text='EXCEL-DATEI BEARBEITEN')
        self.label2.grid(row=0, column=0, sticky='NSEW')

        self.tb2 = scrolledtext.ScrolledText(self.tab2, wrap=tk.WORD, height=25)
        self.tb2.grid(row=1, column=0, columnspan=39, sticky='NSEW')

        self.btn_excel_t2 = ttk.Button(self.tab2, text='Excel-Check starten', command=dummy, width=BWIDTH)
        self.btn_excel_t2.grid(row=29, column=0, padx=2, pady=2)

        # self.btn_report = ttk.Button(self.tab2, text='Report speichern', state=tk.DISABLED, command=dummy,
        #                              width=BWIDTH)
        # self.btn_report.grid(row=29, column=15, padx=2, pady=2)

        self.btn_quit_t2 = ttk.Button(self.tab2, text='Beenden', command=self.click_beenden, width=BWIDTH)
        self.btn_quit_t2.grid(row=29, column=30, padx=2, pady=2)

        # Tab 3 - Details
        self.label3 = ttk.Label(self.tab3, text='DETAILS')
        self.label3.grid(row=0, column=0, sticky='NSEW')

        self.tb3 = scrolledtext.ScrolledText(self.tab3, wrap=tk.WORD, height=25)
        self.tb3.grid(row=1, column=0, columnspan=39, sticky='NSEW')

        self.btn_report_oeffnen = ttk.Button(self.tab3, text='Excel-Report öffnen', default='active',
                                             command=self.starte_excel_report, width=BWIDTH)
        self.btn_report_oeffnen.grid(row=29, column=0, padx=2, pady=2)

        self.btn_quit_t3 = ttk.Button(self.tab3, text='Beenden', command=self.click_beenden, width=BWIDTH)
        self.btn_quit_t3.grid(row=29, column=30, padx=2, pady=2)


        # Tab 4 - Dateiliste
        # todo Duplikate und leere Zeilen löschen
        self.label4 = ttk.Label(self.tab4, text='EINGELESENE DATEIEN')
        self.label4.grid(row=0, column=0)

        self.tb4 = scrolledtext.ScrolledText(self.tab4, relief=tk.SUNKEN, wrap=tk.WORD, height=25)
        self.tb4.grid(row=1, column=0, columnspan=39, sticky='NSEW')
        self.tb4.insert(tk.END, 'Hier stehen nach der Auswahl die Dateien (inkl. Pfade) ...')

        self.btn_waehlen = ttk.Button(self.tab4, text='Datei wählen',
                                      default='active', command=self.click_datei_waehlen, width=BWIDTH)
        self.btn_waehlen.grid(row=29, column=0, padx=2, pady=2)

        self.btn_speichern = ttk.Button(self.tab4, text='Liste speichern',
                                        command=self.click_speicher_dateinamen, width=BWIDTH)
        self.btn_speichern.grid(row=29, column=15, padx=2, pady=2)

        self.btn_quit_t4 = ttk.Button(self.tab4, text='Beenden', command=self.click_beenden, width=BWIDTH)
        self.btn_quit_t4.grid(row=29, column=30, padx=2, pady=2)

        self.update()
        # Tab - Control
        self.existiert_dateiliste()
        return

    def existiert_dateiliste(self):
        self.tb1.delete(1.0, tk.END)
        if os.path.exists('Dateiliste.txt') == True:
            msg = 'Die Datei "Dateiliste.txt" existiert.\n\n'
            print(msg)
            self.tb1.insert(tk.END, msg)
            self.dateiliste_einlesen()
        else:
            msg = 'Die Datei "Dateiliste.txt" existiert N I C H T !!!.\n' \
                  'Du musst die "Dateiliste" im Tab Dateiliste neu erstellen!\n\n'
            print(msg)
            self.tb1.insert(tk.END, msg)
        return

    def click_datei_waehlen(self, event=None):
        # Checken, ob es schon eine Dateiliste gibt und erfolgreich eingelesen wurde
        def lese_datei_ein():
            filename = askopenfilename()
            self.dateiliste.append(filename)
            self.gespeichert = False
            self.fuelle_tb4()
            return
        if os.path.exists('Dateiliste.txt'):
            titel = '"Dateiliste.txt" existiert'
            ergebnis = messagebox.askyesno(titel, 'Es existiert bereits eine Dateiliste! Löschen und neue anlegen?')
            if ergebnis:
                os.remove('Dateiliste.txt')
                self.dateiliste = []
                lese_datei_ein()
        else:
            lese_datei_ein()
        self.tb4.focus()
        return

    def click_speicher_dateinamen(self, event=None):
        # Liste bereinigen: Duplikate löschen und leere Zeilen
        if len(self.dateiliste) > 0:
            self.liste_bereinigen()
            self.fuelle_tb4()
            dateiname = 'Dateiliste.txt'
            def schreibe_datei(dname, dliste):
                with open(dname, 'w') as writefile:
                    for line in dliste:
                        writefile.write(line + '\n')
                self.gespeichert = True
                return
            if os.path.isfile(dateiname):
                ergebnis = messagebox.askyesno("Datei existiert!",
                                               "Die Datei existiert bereits! Wollen Sie sie überschreiben?")
                if ergebnis:
                    schreibe_datei(dateiname, self.dateiliste)
                else:
                    return
            else:
                schreibe_datei(dateiname, self.dateiliste)

        else:
            messagebox.showinfo("Speichern?", "Gibt nichts zu speichern!")
            self.tab4.focus()
        return

    def liste_bereinigen(self):
        if len(self.dateiliste) > 0:
            # Das ist ein Trick, die Liste in ein Set umzuwandeln, denn Sets haben keine Duplikate und dann
            # wieder zurück in eine Liste
            self.dateiliste = list(set(self.dateiliste))
            self.dateiliste.sort()
            loesch_index = []
            for ind, line in enumerate(self.dateiliste):
                if line == '':
                    loesch_index.append(ind)
            for i in sorted(loesch_index, reverse=True):
                del loesch_index[i]
        return

    def click_beenden(self):
        if self.gespeichert:
            self.fenster.destroy()
            sys.exit()
        else:
            ergebnis = messagebox.askyesno('Speichern?',
                                           'Die "Dateiliste.txt" wurde noch nicht gespeichert! Wirklich beenden?')
            if ergebnis:
                self.fenster.destroy()
                sys.exit()
            else:
                self.tab4.focus()
        return

    def excel_bearbeiten(self):
        # todo wenn xlsb, konvertiere zu xlsx
        # Solange die Bearbeitung läuft, wird der Button disabled
        self.btn_excel_t2.configure(state='disabled')
        #Liste_Fehler = []
        msg = 'Die Dateiliste enthält folgende Dateien:\n'
        for line in self.dateiliste:
            msg = msg + line + '\n'
        self.tb2.insert(tk.END, msg+'\n')

        # Alle Dateien nach einander abarbeiten
        for datei in self.dateiliste:
            self.Liste_Fehler.append(Excel_Datei(datei, self))

        self.btn_excel_t2.configure(state='enable')
        self.update()
        return

    def excel_focus(self):
        self.tab_control.select(self.tab2)
        self.update()
        self.excel_bearbeiten()
        return

    def fuelle_tb4(self):
        self.tb4.delete(1.0, tk.END)
        for zeile in self.dateiliste:
            self.tb4.insert(tk.END, zeile + '\n')
        return

    def dateiliste_einlesen(self):
        tmp_dateiliste = []
        try:
            with open('Dateiliste.txt', 'r') as file:
                for line in file:
                    tmp_dateiliste.append(line.strip())
            self.dateiliste = tmp_dateiliste
            msg = f'Die Dateiliste wurde erfolgreich eingelesen. {len(self.dateiliste)} Dateinamen gefunden:\n'
            self.tb1.insert(tk.END, msg)
            self.aktiviere_deaktiviere_elemente(self.tab2, 'enable')
            self.aktiviere_deaktiviere_elemente(self.tab3, 'enable')
            for zeile in self.dateiliste:
                self.tb1.insert(tk.END, zeile+'\n')
            self.checke_alle_dateien()
        except IOError:
            msg = 'FEHLER beim Lesen der Datei! Erstelle eine neue "Dateiliste.txt"!'
            print(msg)
            self.tb1.insert(tk.END, msg)
            self.aktiviere_deaktiviere_elemente(self.tab2, 'disabled')
            self.aktiviere_deaktiviere_elemente(self.tab3, 'disabled')
        return

    def checke_alle_dateien(self):
        """
        Diese Funktion testet ob alle Dateien in der Datei-Liste an der Stelle auch tatsächlich existieren
        :return: kein Rückgabewert
        """
        self.fehler = False
        for ind, datei in enumerate(self.dateiliste):
            if os.path.exists(datei):
                msg = f'\nDie {ind+1}. Datei "{datei}" existiert.\n'
                print(msg)
                self.tb1.insert(tk.END, msg)
            else:
                msg = f'\nDie {ind+1}. Datei "{datei}" existiert N I C H T !!!.\n'
                print(msg)
                self.tb1.insert(tk.END, msg)
                self.fehler = True
        if self.fehler:
            # Elemente deaktivieren
            msg = '\nDa es ein Problem mit einer Datei in der "Dateiliste.txt" gibt, sind einige Elemente ' \
                  'deaktiviert. Du musst die "Dateiliste.txt" in dem Tab Dateiliste neu erstellen.\n'
            print(msg)
            self.tb1.insert(tk.END, msg)
            self.aktiviere_deaktiviere_elemente(self.tab2, 'disabled')
            self.aktiviere_deaktiviere_elemente(self.tab3, 'disabled')
            self.btn_excel.configure(state='disabled')
            self.btn_quit_t1.configure(default='active')
        return

    def aktiviere_deaktiviere_elemente(self, element, state):
        for ele in element.winfo_children():
            try:
                ele.configure(state=state)
            except:
                pass
        #self.btn_report.configure(state=tk.DISABLED)
        return

    def konvertiere_xlsb_datei(self, inp_pfad, out_pfad):
        from win32com.client import Dispatch
        """
        Diese Funktion wandelt einen Excel-XLSB-Datei in eine XLSX-Datei um
        Input: String für input datei und output datei
        """
        # Starte Excel
        excel = Dispatch("Excel.Application")  # excel ist in diesem Fall wirklich nur ein Variablen-Name
        # Schalte Alle Warnungen aus, z.B. dass es noch verlinkte Dateien gibt, oder das beim speichern die Makros
        # verloren gehen
        excel.DisplayAlerts = False
        # Excel ist sichtbar auf dem Desktop
        excel.Visible = True
        # Öffne die entsprechende (inp_pfad) Datei
        wb = excel.Workbooks.Open(inp_pfad)
        # Speichere im xlsx-format
        wb.SaveAs(out_pfad, FileFormat=51)
        # Datei schliessen
        wb.Close()
        # Excel schliessen
        excel.Quit()
        return

    def starte_excel_report(self):
        if len(self.Liste_Fehler) > 0:
            from win32com.client import Dispatch
            """
            Diese Funktion öffnet den erstellten Excel-Report
            Input: String für input datei
            """
            # Starte Excel
            excel = Dispatch("Excel.Application")  # excel ist in diesem Fall wirklich nur ein Variablen-Name
            # Schalte Alle Warnungen aus, z.B. dass es noch verlinkte Dateien gibt, oder das beim speichern die Makros
            # verloren gehen
            excel.DisplayAlerts = False
            # Excel ist sichtbar auf dem Desktop
            excel.Visible = True
            # Öffne die entsprechende (inp_pfad) Datei
            wb = excel.Workbooks.Open(self.Liste_Fehler[-1].report_dateiname)
        else:
            messagebox.showwarning('Zu früh!', 'Erst muss ein Report erstellt werden, bevor man ihn zeigen kann! :)')
        return


class Excel_Datei(object):
    def __init__(self, datei, root):
        self.dateiname = datei
        self.formelspalten = []
        self.fehlerliste = {}
        self.formel_fehlt = {}
        self.letzte_zeile = 0
        self.anz_zellen = 0
        self.anz_fehler = 0
        self.anz_fehler_spalten = 0
        self.root = root
        self.platform = ''
        self.umwandlung_moeglich = False
        self.datei_typ = ''
        self.report_erstellt = True
        self.report_dateiname = ''

        # Checke die Platform auf dem das Programm ausgeführt wird.
        self.check_platform()

        self.root.tb2.delete('1.0', tk.END)
        self.root.update()

        # Checke ob Datei XLSX oder XLSB ist -> ggf. Umwandlung nötig
        self.datei_typ = self.checke_datei_typ(self.dateiname)
        if self.datei_typ == 'xlsx':
            msg = f'Die Datei {self.dateiname} wird geladen. GEDULD!\n\n'
            print(msg)
            self.root.tb2.insert(tk.END, msg)
            self.root.update()
            self.lade_xlsx(self.dateiname)
        elif self.datei_typ == 'xlsb':
            # Konvertierung möglich
            msg = 'Konvertierung NÖTIG!\n\n'
            print(msg)
            self.root.tb2.insert(tk.END, msg)
            self.root.update()
            self.root.click_beenden()
        else:
            print('Dateiformat nicht erkannt!')
            self.root.click_beenden()

        msg = 'Die Formelspalten werden identifiziert ...\n'
        print(msg)
        self.root.tb2.insert(tk.END, msg)
        self.root.update()
        self.finde_spalten_mit_formeln(self.arbeitsblatt_formel)

        msg = 'Die letzte relevante Zeile (=Musterzeile) wird identifiziert ...\n'
        print(msg)
        self.root.tb2.insert(tk.END, msg)
        self.root.update()
        self.finde_letzte_rel_zeile(self.arbeitsblatt_formel)

        msg = 'Jetzt werden alle Formel-Spalten identifiziert, in denen Fehler sind ...\n\n'
        print(msg)
        self.root.tb2.insert(tk.END, msg)
        self.root.update()
        self.finde_fehler_zellen(self.arbeitsblatt_data)
        self.ermittel_anz_fehler()

        msg = '\nIch durchsuche alle "Soll"-Formelspalten auf fehlende Formeln  ...\n\n'
        print(msg)
        self.root.tb2.insert(tk.END, msg)
        self.root.update()
        self.finde_zellen_ohne_formel(self.arbeitsblatt_formel)

        # Fuelle Tab3 mit Details
        self.fuelle_detail_tab()
        self.erstelle_excel_report()

        # Nach dem Schreiben des Report "lösche" ich die eingelesenen Excel-Files
        self.arbeitsblatt_formel = None
        self.arbeitsblatt_data = None
        return

    def fuelle_detail_tab(self):
        msg = f'Datei: {self.dateiname}\n'
        msg += f'Anzahl Formelspalten: {len(self.formelspalten)}\n'
        msg += f'Letzte relevante (untersuchte) Zeile: {self.letzte_zeile}\n'
        msg += f'Anzahl untersuchter Zellen: {self.anz_zellen}\n'
        msg += f'Anzahl verschiedener Spalten mit Fehlern: {self.anz_fehler_spalten}\n'
        msg += f'Anzahl gefundener Fehler: {self.anz_fehler}\n'
        msg += f'Anzahl "Soll"-Formelspalten in denen Formeln fehlen: {len(self.formel_fehlt)}\n'
        self.root.tb3.insert(tk.END, msg)
        self.root.update()

        # Alle Zellen ausgeben, in denen Fehler sind
        if len(self.formel_fehlt) > 0:
            for key in self.fehlerliste:
                msg = f'In der Spalte {key} sind {len(self.fehlerliste[key])} Fehler.\n'
                self.root.tb3.insert(tk.END, msg)
                self.root.update()
            for key in self.fehlerliste:
                for item in self.fehlerliste[key]:
                    msg = f'Fehler in: {key+str(item)}\n'
                    self.root.tb3.insert(tk.END, msg)

        # Alle Zellen ausgeben in denen eine Formel sein sollte, aber nicht ist
        if len(self.formel_fehlt) > 0:
            for key in self.formel_fehlt:
                msg = f'In der Spalte {key} fehlen {len(self.formel_fehlt[key])} Formeln.\n'
                self.root.tb3.insert(tk.END, msg)
                self.root.update()
            for key in self.formel_fehlt:
                for item in self.formel_fehlt[key]:
                    msg = f'Formel sollte sein, aber ist nicht in: {key+str(item)}\n'
                    self.root.tb3.insert(tk.END, msg)

        msg = (28*'+-') +'\n'
        self.root.tb3.insert(tk.END, msg)
        self.root.update()
        return

    def check_platform(self):
        import platform
        my_sys = platform.system()
        if ('Darwin' in my_sys) or ('Linux' in my_sys):
            # Das Skript läuft unter Mac oder Linux
            msg = 'Das Programm läuft unter Mac/Linux, es wird kein Check gemacht, ob Excel läuft.\n'
            msg += 'XLSB - Dateien können NICHT umgewandelt werden.\n\n'
            print(msg)
            self.root.tb2.insert(tk.END, msg + '\n')
            self.platform = my_sys
            self.umwandlung_moeglich = False
            return
        elif 'Windows' in my_sys:
            import win32com.client
            import pythoncom
            try:
                win32com.client.GetActiveObject("Excel.Application")
                # Wenn es bei dem Befehl oben keinen Fehler gibt, dann läuft Excel
                msg = 'MS EXCEL läuft! Beenden Sie Excel und starten Sie dann dieses Programm neu!\n\n'
                self.root.tb2.insert(tk.END, msg)
                messagebox.showerror('FEHLER!', msg)
                self.root.click_beenden()
            except pythoncom.com_error as error:
                print('Excel is NOT running, this is good!')
                self.platform = my_sys
                self.umwandlung_moeglich = True
        else:
            msg = 'PROBLEM: Ich habe dieses System nicht erkannt!\n'
            msg += f'Erkannt wurde: {my_sys}! Das Programm wird beendet. Wenden Sie sich an den Admin!'
            print(msg)
            self.root.tb2.insert(tk.END, msg + '\n')
            messagebox.showerror('Fehler', msg)
            self.root.click_beenden()
        return

    def checke_datei_typ(self, datei):
        if datei[-4:] == 'xlsx':
            # XLSX - Datei kann direkt verarbeitet werden, keine Umwandlung nötig
            msg = datei + ' -- XLSX-Datei identifiziert!\n'
            self.root.tb2.insert(tk.END, msg + '\n')
            return 'xlsx'
        elif datei[-4:] == 'xlsb':
            # XLSB - Datei! Erst in XLSX umwandeln und dann wie XLSX
            msg = datei + ' -- XLSB-Datei identifiziert!\n'
            msg += 'Diese Datei muss umgewandelt werden ... \n'
            msg += 'Die Umwandlung kann nur unter Windows mit installiertem Excel statt finden.\n'
            self.root.tb2.insert(tk.END, msg + '\n')
            return 'xlsb'
        else:
            # Keine weitere Verarbeitung !
            msg = datei + ' -- Keine XLSX oder XLSB-Datei identifiziert! Diese Datei wird übersprungen!\n'
            self.root.tb2.insert(tk.END, msg + '\n')
        return

    def lade_xlsx(self, datei):
        from openpyxl import load_workbook
        wb_formula = load_workbook(datei, data_only=False)
        wb_data = load_workbook(datei, data_only=True)
        self.arbeitsblatt_formel = wb_formula['Warenwirtschaft']
        self.arbeitsblatt_data = wb_data['Warenwirtschaft']
        return

    def finde_spalten_mit_formeln(self, arbeitsblatt):
        """
        Diese Funktion ermittelt in einem Arbeitsblatt alle Spalten in denen sich Formeln befinden.
        Massgeblich ist hier die Zeile 5 - die erste Zeile, in der die Werte stehen.
        Fehlt in dieser Zeile 5 bereits eine Formel, wird die Spalte von dieser Funktion auch ignoriert /
        bzw. nicht als Formel-Spalte erkannt.
        Input: Ein Arbeitsblatt einer Excel-Tabelle (openpyxl.workbook)
        Output: Liste mit Spaltennummern (Int), in denen sich Formeln befinden
        """
        alle_formeln_spalten = []
        for spalte in range(1, arbeitsblatt.max_column):
            if arbeitsblatt.cell(row=5, column=spalte).data_type == 'f':
                alle_formeln_spalten.append(spalte)
        self.formelspalten = alle_formeln_spalten
        msg = f'Es wurden {len(self.formelspalten)} Spalten mit Formeln identifiziert!\n\n'
        self.root.tb2.insert(tk.END, msg)
        return

    def finde_letzte_rel_zeile(self, arbeitsblatt):
        """
        Diese Funktion identifiziert die letzte relevante Zeile.
        Dafür wird in Spalte 1 jede Zeile angeschaut, bis das Stichwort 'Musterzeile' gefunden wird.
        TBD: Die darüber enthaltenen Leerzeilen, müssen noch identifiziert und abgezogen werden
        Input: Ein Arbeitsblatt einer Excel-Tabelle (openpyxl.workbook)
        Output: Letzte relevante Zeile (Int)
        """
        musterzeile_gefunden_in_zeile = 0
        for row in range(5, arbeitsblatt.max_row+2):
            if arbeitsblatt.cell(row=row, column=1).value is None:
                print(f'Problem mit Zeile: {row} -- {arbeitsblatt.cell(row=row, column=1).value}')
                pass
            else:
                if "Musterzeile" in arbeitsblatt.cell(row=row, column=1).value:
                    musterzeile_gefunden_in_zeile = row
                    print('Musterzeile in', row)
                    break
        for row in range(musterzeile_gefunden_in_zeile-1, 0, -1):
            if arbeitsblatt.cell(row=row, column=1).value == None:
                pass
            else:
                print(arbeitsblatt.cell(row=row, column=1).value, ' --- ', row)
                letzte_relevante_zeile = row
                break
        self.letzte_zeile = letzte_relevante_zeile
        msg = f'Die letzte relevante Zeile ist {self.letzte_zeile}.\n\n'
        self.root.tb2.insert(tk.END, msg)
        return

    def spalte_zu_string(self, column):
        string = ''
        while column > 0:
            column, rest = divmod(column - 1, 26)
            string = chr(65 + rest) + string
        return string

    def finde_fehler_zellen(self, arbeitsblatt):
        """
        Diese Funktion identifiziert jede Zelle, in der sich ein Fehler befindet.
        Das kann #Wert, #Bezug, oder sonst ein Fehler sein.
        Input: Ein Arbeitsblatt einer Excel-Tabelle (openpyxl.workbook), alle relevanten Zeilen (zeilen=int),
        alle Formelspalten (liste mit int)
        Output: Eine Liste mit Fehlerzellen. Tupel (row, column)
        """
        fehlerzellen = {}
        zaehler = 0
        for fs in self.formelspalten:
            for z in range(5, self.letzte_zeile):
                zaehler += 1
                if arbeitsblatt.cell(row=z, column=fs).data_type == 'e':
                    # Fehler gefunden
                    # Checke ob diese Spalte bereits im dict ist
                    if self.spalte_zu_string(fs) in fehlerzellen:
                        fehlerzellen[self.spalte_zu_string(fs)].append(z)
                    else:
                        # befindet sich noch nicht im dict
                        fehlerzellen[self.spalte_zu_string(fs)] = []
                        fehlerzellen[self.spalte_zu_string(fs)].append(z)
                    print(f'Fehlerzelle gefunden! {self.spalte_zu_string(fs)}{z}')
                    try:
                        print(arbeitsblatt.cell(row=row, column=1).value)
                    except:
                        pass
        self.anz_zellen = zaehler
        self.fehlerliste = fehlerzellen
        for key in self.fehlerliste:
            msg = f'SPALTE {key} hat {len(self.fehlerliste[key])} Fehler!\n'
            self.root.tb2.insert(tk.END, msg)
        return

    def finde_zellen_ohne_formel(self, arbeitsblatt):
        # Gehe durch jede Zeile und checke ob die "Soll"-Formelspalte auch eine Formel enthält
        for zeile in range(5, self.letzte_zeile+1):
            for spalte in self.formelspalten:
                if arbeitsblatt.cell(row=zeile, column=spalte).data_type != 'f':
                    if self.spalte_zu_string(spalte) not in self.formel_fehlt:
                        self.formel_fehlt[self.spalte_zu_string(spalte)] = []
                        self.formel_fehlt[self.spalte_zu_string(spalte)].append(zeile)
                    else:
                        self.formel_fehlt[self.spalte_zu_string(spalte)].append(zeile)
        self.report_erstellt = False
        return

    def ermittel_anz_fehler(self):
        if len(self.fehlerliste) > 0:
            zaehler = 0
            for key in self.fehlerliste:
                zaehler += len(self.fehlerliste[key])
            self.anz_fehler = zaehler
            self.anz_fehler_spalten = len(self.fehlerliste)
        return

    def neuer_report_name(self, dname):
        index = 0
        for ind, c in enumerate(dname[::-1]):
            if c == '/':
                index = ind
                break
        neuer_dname = f'{dname[:-index]}{str(time.strftime("%Y%m%d_%H%M%S"))}_report_{dname[-index:]}'
        return neuer_dname, dname[-index:]

    def erstelle_excel_report(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, colors
        wb = Workbook()

        self.report_dateiname, sheet_name = self.neuer_report_name(self.dateiname)

        ws1 = wb.create_sheet(0)
        ws1.title = sheet_name

        ft1 = Font(color=colors.BLACK, bold=True, size=22)
        ft2 = Font(color=colors.BLACK, size=14)
        ft3 = Font(color=colors.BLACK, bold=True, size=16)
        ft4 = Font(color=colors.BLUE, underline='single', size=12)

        dateiname = f'{self.dateiname}#Warenwirtschaft'
        to_del = wb.get_sheet_by_name('Sheet')

        start_zeile = 11
        for key in self.fehlerliste:
            for item in self.fehlerliste[key]:
                link_text = f'Fehler in Formel in Zelle {key+str(item)}'
                link = f'=HYPERLINK("File:///{dateiname}!{key+str(item)}","{link_text}")'
                ws1['A' + str(start_zeile)] = link
                ws1['A' + str(start_zeile)].font = ft4
                start_zeile += 1

        start_zeile = 11
        for key in self.formel_fehlt:
            for item in self.formel_fehlt[key]:
                link_text = f'Fehlende Formel in Zelle {key+str(item)}'
                link = f'=HYPERLINK("File:///{dateiname}!{key+str(item)}","{link_text}")'
                ws1['C' + str(start_zeile)] = link
                ws1['C' + str(start_zeile)].font = ft4
                start_zeile += 1

        # # kopiert von Stackoverflow: Passt die Spaltenbreite an.
        # for column_cells in ws1.columns:
        #     length = max(len(str(cell.value)) for cell in column_cells)
        #     ws1.column_dimensions[column_cells[0].column].width = length

        ws1.column_dimensions['A'].width = 38
        ws1.column_dimensions['C'].width = 38

        ws1['A1'] = 'Report: ' + self.dateiname
        ws1['A1'].font = ft1

        ws1['A3'] = f'Anzahl Formelspalten: {locale.format("%.0f", len(self.formelspalten), True)}'
        ws1['A3'].font = ft2

        ws1['A4'] = f'Letzte relevante (untersuchte) Zeile: {locale.format("%.0f", self.letzte_zeile, True)}'
        ws1['A4'].font = ft2

        ws1['A5'] = f'Anzahl untersuchter Zellen: {locale.format("%.0f", self.anz_zellen, True)}'
        ws1['A5'].font = ft2

        ws1['A6'] = f'Anzahl verschiedener Spalten mit Fehlern: {locale.format("%.0f", self.anz_fehler_spalten, True)}'
        ws1['A6'].font = ft2

        ws1['A7'] = f'Anzahl gefundener Fehler: {locale.format("%.0f", self.anz_fehler, True)}'
        ws1['A7'].font = ft2

        ws1['A8'] = f'Anzahl "Soll"-Formelspalten in denen Formeln fehlen: ' \
                    f'{locale.format("%.0f", len(self.formel_fehlt), True)}'
        ws1['A8'].font = ft2

        ws1['A10'] = 'Fehler in Formeln'
        ws1['A10'].font = ft3

        ws1['C10'] = 'Fehlende Formeln'
        ws1['C10'].font = ft3

        ws1.merge_cells('A1:S1')
        ws1.merge_cells('A3:C3')
        ws1.merge_cells('A4:C4')
        ws1.merge_cells('A5:C5')
        ws1.merge_cells('A6:C6')
        ws1.merge_cells('A7:C7')
        ws1.merge_cells('A8:C8')

        wb.remove_sheet(to_del)
        msg = f'Speichere den Excelreport unter: {self.report_dateiname}!\n'
        self.root.tb2.insert(tk.END, msg)
        wb.save(self.report_dateiname)
        return

def dummy():
    print('Do nothing!')
    return

if __name__ == "__main__":
    root = tk.Tk()
    rows = 0
    while rows < 30:
        root.rowconfigure(rows, weight=1)
        root.columnconfigure(rows, weight=1)
        rows += 1
    Klaus_App(root).grid(sticky='NSEW')
    root.mainloop()